import os, fnmatch
import openpyxl

endnum=90

def find(pattern):
    result = []
    for root, dirs, files in os.walk(os.getcwd()):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                return  name
                # print name
                # result.append(os.path.realpath(name))


ctr=50

# for ctr in range(ctr,endnum+10,10):
file1 = str(ctr) + '*.xlsx'
filename=find(file1)
print filename

wb=openpyxl.load_workbook(filename)
ws=wb.active
sheetname=wb.sheetnames
print sheetname

cellname=ws.cell(row=5, column=2).value
print str(cellname)