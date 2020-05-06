import openpyxl
from openpyxl import Workbook
import os, fnmatch

startnum=input("Starting value ")
endnum=input("Ending value ")
chnum=input("Select channel ")+1
skip_dist=input("Enter skip distance ")
ctr=startnum

def find(pattern):
    result = []
    for root, dirs, files in os.walk(os.getcwd()):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                return  name

# for k in range(2,Channelcount+3, 1):
combined=Workbook()
Comb_sheet=combined.active

# copy frequency column
file1 = str(startnum) + '*.xlsx'
filename = find(file1)
wb1=openpyxl.load_workbook(filename)
ws1 = wb1.active

rowsel=[]
for i in range(4,38,1):
    rowsel.append(ws1.cell(row = i,column = 1).value)

countrow=0
for i in range(4,38,1):
    Comb_sheet.cell(row = i,column = 1).value=rowsel[countrow]
    countrow +=1


#copy the rest

startrange=2
endrange=startrange+((endnum-startnum)/10)+1
for j in range(startrange,endrange,skip_dist/10):
    file1 = str(ctr) + '*.xlsx'
    filename = find(file1)
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.active
    rowsel = []
    for i in range(4, 38, 1):
        rowsel.append(ws1.cell(row=i, column=chnum).value)

    countrow = 0
    for i in range(4, 38, 1):
        Comb_sheet.cell(row=3, column=j).value=str(wb1.sheetnames)
        Comb_sheet.cell(row=i, column=j).value = rowsel[countrow]
        countrow += 1

    ctr += skip_dist

savefilename=str(ws1.cell(row=5, column=chnum).value)+'.xlsx'
combined.save(savefilename)

